import logging
import uuid
from pathlib import Path
from core import ollama_client, rag_service
from documents.prompts import DOCUMENT_SUMMARY_PROMPT

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Extracts text from documents, summarizes, and indexes in ChromaDB."""

    async def process(self, file_path: str, file_name: str, file_type: str, user_id: int) -> dict:
        """Process a document: extract text, summarize, index."""
        doc_id = str(uuid.uuid4())

        # Extract text
        text, page_count = self._extract_text(file_path, file_type)
        if not text.strip():
            return {
                "document_id": doc_id,
                "file_name": file_name,
                "summary": "Could not extract text from this document.",
                "entities": [],
                "page_count": 0,
                "indexed": False,
            }

        # Summarize with LLM
        summary, entities = await self._summarize(text, file_name)

        # Index in ChromaDB
        indexed = False
        try:
            rag_service.add_document(
                text=text,
                metadata={
                    "file_name": file_name,
                    "file_type": file_type,
                    "user_id": user_id,
                    "summary": summary[:500],
                },
                doc_id=doc_id,
            )
            indexed = True
        except Exception as e:
            logger.error(f"Failed to index document: {e}")

        return {
            "document_id": doc_id,
            "file_name": file_name,
            "summary": summary,
            "entities": entities,
            "page_count": page_count,
            "indexed": indexed,
        }

    def _extract_text(self, file_path: str, file_type: str) -> tuple[str, int]:
        """Extract text content from a file."""
        extractors = {
            "pdf": self._extract_pdf,
            "docx": self._extract_docx,
            "xlsx": self._extract_xlsx,
            "csv": self._extract_csv,
            "txt": self._extract_txt,
        }
        extractor = extractors.get(file_type)
        if not extractor:
            return "", 0
        try:
            return extractor(file_path)
        except Exception as e:
            logger.error(f"Text extraction failed for {file_path}: {e}")
            return "", 0

    @staticmethod
    def _extract_pdf(file_path: str) -> tuple[str, int]:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages), len(reader.pages)

    @staticmethod
    def _extract_docx(file_path: str) -> tuple[str, int]:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs), len(paragraphs) // 25 + 1

    @staticmethod
    def _extract_xlsx(file_path: str) -> tuple[str, int]:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sheets_text.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets_text), len(wb.sheetnames)

    @staticmethod
    def _extract_csv(file_path: str) -> tuple[str, int]:
        text = Path(file_path).read_text(encoding="utf-8", errors="ignore")
        lines = text.strip().split("\n")
        return text, max(1, len(lines) // 50)

    @staticmethod
    def _extract_txt(file_path: str) -> tuple[str, int]:
        text = Path(file_path).read_text(encoding="utf-8", errors="ignore")
        return text, 1

    async def _summarize(self, text: str, file_name: str) -> tuple[str, list[str]]:
        """Summarize document and extract entities using LLM."""
        # Truncate very long documents
        truncated = text[:3000] if len(text) > 3000 else text

        prompt = DOCUMENT_SUMMARY_PROMPT.format(
            file_name=file_name,
            content=truncated,
        )

        try:
            result = await ollama_client.generate(prompt, temperature=0.3)

            # Try to extract entities from the response
            entities = []
            lines = result.split("\n")
            in_entities = False
            for line in lines:
                line = line.strip()
                if "entities" in line.lower() or "key entities" in line.lower():
                    in_entities = True
                    continue
                if in_entities and line.startswith("-"):
                    entity = line.lstrip("- ").strip()
                    if entity:
                        entities.append(entity)
                elif in_entities and not line:
                    in_entities = False

            return result, entities[:20]
        except Exception as e:
            logger.error(f"Document summarization failed: {e}")
            return f"Document '{file_name}' processed but summarization is unavailable.", []


document_processor = DocumentProcessor()
